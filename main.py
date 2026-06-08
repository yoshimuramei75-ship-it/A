from fastapi import FastAPI, UploadFile, File
from fastapi.responses import HTMLResponse, StreamingResponse
from pydantic import BaseModel
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import os
import uvicorn
import io
import urllib.parse
import json
from typing import List, Optional
from contextlib import asynccontextmanager

@asynccontextmanager
async def lifespan(app: FastAPI):
    print("\n" + "="*50)
    if os.path.exists("日報改善.xlsx"):
        print(f"✅ Excelファイルを確認しました")
    else:
        print(f"❌ ERROR: '日報改善.xlsx' がありません")
    print("="*50 + "\n")
    yield

app = FastAPI(lifespan=lifespan)

EXCEL_FILE = "日報改善.xlsx"
SYNC_FILE = "data_sync.json"

def read_master(wb, sheet_name, max_col):
    data = []
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, max_col=max_col, values_only=True):
            clean_row = [str(c).strip() if c is not None else "" for c in row]
            if any(clean_row):
                data.append(clean_row)
    return data

# =========================================================
# 🏠 画面表示用のURL
# =========================================================
@app.get("/", response_class=HTMLResponse)
async def read_index():
    with open("index.html", "r", encoding="utf-8") as f:
        return f.read()

@app.get("/tanka", response_class=HTMLResponse)
async def read_tanka():
    with open("tanka.html", "r", encoding="utf-8") as f:
        return f.read()

# =========================================================
# 🐾 日報アプリ用のAPI
# =========================================================

@app.get("/api/init_data")
async def get_init_data():
    if not os.path.exists(EXCEL_FILE):
        return {"projects": [], "staff": [], "inspectors": ["嶺岸", "熊谷", "江連", "発注者", "検査者"]}

    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

    staff_list = []
    if "Master_Staff" in wb.sheetnames:
        for row in wb["Master_Staff"].iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0]: staff_list.append(str(row[0]).strip())

    project_list = []
    if "Master_Projects" in wb.sheetnames:
        for row in wb["Master_Projects"].iter_rows(min_row=2, max_col=15, values_only=True):
            if row[1]:
                status = row[12] if len(row) >= 13 and row[12] else "稼働現場"
                manager = str(row[5]) if len(row) >= 6 and row[5] else ""
                project_list.append({"no": str(row[1]), "name": str(row[2]), "status": str(status), "manager": manager})

    return {
        "projects": project_list,
        "staff": staff_list,
        "inspectors": ["嶺岸", "熊谷", "江連", "発注者", "検査者"],
        "progress": read_master(wb, "Master_Progress", 2),
        "company": read_master(wb, "Master_Company", 2),
        "role": read_master(wb, "Master_Role", 2),
        "machine": read_master(wb, "Master_Machine", 4),
        "material": read_master(wb, "Master_Material", 8)
    }

class ProjectData(BaseModel):
    project_no: str
    project_name: str
    manager: str
    status: str

@app.post("/api/add_project")
async def add_project(data: ProjectData):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    if "Master_Projects" not in wb.sheetnames:
        wb.create_sheet("Master_Projects")
    ws = wb["Master_Projects"]
    new_row = ["", data.project_no, data.project_name, "", "", data.manager, "", "", "", "", "", "", data.status]
    ws.append(new_row)
    wb.save(EXCEL_FILE)
    return {"status": "success"}

class UpdateStatusRequest(BaseModel):
    project_no: str
    new_status: str

@app.post("/api/update_status")
async def update_status(data: UpdateStatusRequest):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Master_Projects"]
    updated = False
    for row in ws.iter_rows(min_row=2):
        if str(row[1].value) == data.project_no:
            row[12].value = data.new_status
            updated = True
    if updated:
        wb.save(EXCEL_FILE)
    return {"status": "success"}

class KeyData(BaseModel):
    key: str
    value: str

class SyncDataPayload(BaseModel):
    storage: dict

@app.post("/api/push_key")
async def push_key(data: KeyData):
    server_data = {}
    if os.path.exists(SYNC_FILE):
        try:
            with open(SYNC_FILE, "r", encoding="utf-8") as f: server_data = json.load(f)
        except: pass
    server_data[data.key] = data.value
    with open(SYNC_FILE, "w", encoding="utf-8") as f: json.dump(server_data, f, ensure_ascii=False)
    return {"status": "success"}

@app.post("/api/push_all")
async def push_all(data: SyncDataPayload):
    server_data = {}
    if os.path.exists(SYNC_FILE):
        try:
            with open(SYNC_FILE, "r", encoding="utf-8") as f: server_data = json.load(f)
        except: pass
    for k, v in data.storage.items():
        server_data[k] = v
    with open(SYNC_FILE, "w", encoding="utf-8") as f: json.dump(server_data, f, ensure_ascii=False)
    return {"status": "success"}

@app.get("/api/pull_all")
async def pull_all():
    if os.path.exists(SYNC_FILE):
        try:
            with open(SYNC_FILE, "r", encoding="utf-8") as f: return {"status": "success", "data": json.load(f)}
        except: pass
    return {"status": "success", "data": {}}

class ReportData(BaseModel):
    date: str
    projectName: str
    manager: str
    temp: str
    weather: str
    am_time: str
    am_abnormal: str
    am_text: str
    pm_time: str
    pm_abnormal: str
    pm_text: str
    inspection_type: Optional[str] = ""
    inspector: str
    judge: Optional[str] = ""
    tomorrow_work: str
    customer: str
    internal: str
    safety: str
    work_data: List[List[str]] = []
    labor_data: List[List[str]] = []
    machine_data: List[List[str]] = []
    material_data: List[List[str]] = []
    labor_total_project: float = 0.0
    machine_total_project: float = 0.0
    stat_days: str = "0"
    stat_hours: str = "0"

class ExportExcelRequest(BaseModel):
    project_no: str
    period: str
    reports: List[ReportData]

def parse_num(val):
    try:
        v = float(val)
        return int(v) if v.is_integer() else v
    except:
        return 0

def format_num(val):
    return int(val) if float(val).is_integer() else float(val)

@app.post("/api/export_excel")
async def export_excel(req: ExportExcelRequest):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    font_bold = Font(bold=True)
    font_title = Font(size=18, bold=True)
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    align_top_left = Alignment(horizontal='left', vertical='top', wrap_text=True)
    thin = Side(style='thin')
    fill_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    def apply_border_to_row(ws, r, c_start, c_end):
        for c in range(c_start, c_end + 1):
            ws.cell(row=r, column=c).border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for report in req.reports:
        ws = wb.create_sheet(title=report.date.replace('-', ''))

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15

        ws.merge_cells('A1:D2')
        ws['A1'] = "作 業 日 報"
        ws['A1'].font = font_title
        ws['A1'].alignment = align_center
        apply_border_to_row(ws, 1, 1, 4)
        apply_border_to_row(ws, 2, 1, 4)

        ws['A3'] = "工事名:"
        ws['B3'] = report.projectName
        ws['C3'] = "日付:"
        ws['D3'] = report.date

        ws['A4'] = "天候/気温:"
        ws['B4'] = f"{report.weather} / {report.temp}"
        ws['C4'] = "作業所長:"
        ws['D4'] = report.manager

        ws['A5'] = "工事日数:"
        ws['B5'] = f"{report.stat_days} 日目"
        ws['C5'] = "無災害時間:"
        ws['D5'] = f"{report.stat_hours} 時間"

        for r in [3, 4, 5]:
            ws[f'A{r}'].font = font_bold; ws[f'A{r}'].fill = fill_gray; ws[f'A{r}'].alignment = align_center
            ws[f'C{r}'].font = font_bold; ws[f'C{r}'].fill = fill_gray; ws[f'C{r}'].alignment = align_center
            apply_border_to_row(ws, r, 1, 4)

        row = 7

        ws[f'A{row}'] = "【本日の作業】"
        ws[f'A{row}'].font = Font(bold=True, size=11, color="673AB7")
        row += 1
        ws[f'A{row}'] = "種別"; ws[f'B{row}'] = "施工箇所"; ws.merge_cells(f'C{row}:D{row}'); ws[f'C{row}'] = "作業内容"
        for c in range(1, 5):
            ws.cell(row=row, column=c).fill = fill_gray; ws.cell(row=row, column=c).font = font_bold; ws.cell(row=row, column=c).alignment = align_center
        apply_border_to_row(ws, row, 1, 4)
        row += 1
        if report.work_data:
            for wd in report.work_data:
                ws[f'A{row}'] = wd[0] if len(wd)>0 else ""; ws[f'B{row}'] = wd[1] if len(wd)>1 else ""; ws.merge_cells(f'C{row}:D{row}'); ws[f'C{row}'] = wd[2] if len(wd)>2 else ""
                apply_border_to_row(ws, row, 1, 4); row += 1
        else:
            ws.merge_cells(f'C{row}:D{row}'); apply_border_to_row(ws, row, 1, 4); row += 1
        row += 1

        ws[f'A{row}'] = "【労務】"
        ws[f'A{row}'].font = Font(bold=True, size=11, color="1A73E8")
        row += 1
        ws.cell(row=row, column=1, value="業者名").fill = fill_gray
        ws.cell(row=row, column=2, value="役割").fill = fill_gray
        ws.cell(row=row, column=3, value="人数").fill = fill_gray
        ws.cell(row=row, column=4, value="").fill = fill_gray
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        for c in range(1, 5):
            ws.cell(row=row, column=c).font = font_bold; ws.cell(row=row, column=c).alignment = align_center
        apply_border_to_row(ws, row, 1, 4); row += 1

        daily_labor = 0
        if report.labor_data:
            for ld in report.labor_data:
                ws.cell(row=row, column=1, value=ld[0])
                ws.cell(row=row, column=2, value=ld[1])
                val_str = ld[2] if len(ld)>2 else "0"
                ws.cell(row=row, column=3, value=val_str)
                ws.cell(row=row, column=4, value="")
                ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
                apply_border_to_row(ws, row, 1, 4)
                daily_labor += parse_num(val_str)
                row += 1
        else:
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
            apply_border_to_row(ws, row, 1, 4); row += 1

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.cell(row=row, column=1, value="本日合計 / 全体累計 (本日まで)").font = font_bold
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='right', vertical='center')
        ws.cell(row=row, column=3, value=f"{format_num(daily_labor)} 人 / {format_num(report.labor_total_project)} 人").font = font_bold
        ws.cell(row=row, column=3).alignment = align_center
        ws.cell(row=row, column=4, value="")
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        apply_border_to_row(ws, row, 1, 4)
        row += 2

        ws[f'A{row}'] = "【重機】"
        ws[f'A{row}'].font = Font(bold=True, size=11, color="0F9D58")
        row += 1
        ws.cell(row=row, column=1, value="機械名").fill = fill_gray
        ws.cell(row=row, column=2, value="規格").fill = fill_gray
        ws.cell(row=row, column=3, value="台数").fill = fill_gray
        ws.cell(row=row, column=4, value="").fill = fill_gray
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        for c in range(1, 5):
            ws.cell(row=row, column=c).font = font_bold; ws.cell(row=row, column=c).alignment = align_center
        apply_border_to_row(ws, row, 1, 4); row += 1

        daily_machine = 0
        if report.machine_data:
            for md in report.machine_data:
                ws.cell(row=row, column=1, value=md[0])
                ws.cell(row=row, column=2, value=md[1])
                val_str = md[2] if len(md)>2 else "0"
                ws.cell(row=row, column=3, value=val_str)
                ws.cell(row=row, column=4, value="")
                ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
                apply_border_to_row(ws, row, 1, 4)
                daily_machine += parse_num(val_str)
                row += 1
        else:
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
            apply_border_to_row(ws, row, 1, 4); row += 1

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.cell(row=row, column=1, value="本日合計 / 全体累計 (本日まで)").font = font_bold
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='right', vertical='center')
        ws.cell(row=row, column=3, value=f"{format_num(daily_machine)} 台 / {format_num(report.machine_total_project)} 台").font = font_bold
        ws.cell(row=row, column=3).alignment = align_center
        ws.cell(row=row, column=4, value="")
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        apply_border_to_row(ws, row, 1, 4)
        row += 2

        ws[f'A{row}'] = "【資材・受入検査】"
        ws[f'A{row}'].font = Font(bold=True, size=11, color="FF9800")
        row += 1
        ws.cell(row=row, column=1, value="資材詳細").fill = fill_gray
        ws.cell(row=row, column=2, value="数量").fill = fill_gray
        ws.cell(row=row, column=3, value="結果").fill = fill_gray
        ws.cell(row=row, column=4, value="").fill = fill_gray
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        for c in range(1, 5):
            ws.cell(row=row, column=c).font = font_bold; ws.cell(row=row, column=c).alignment = align_center
        apply_border_to_row(ws, row, 1, 4); row += 1
        if report.material_data:
            for md in report.material_data:
                ws.cell(row=row, column=1, value=md[0].replace('\n', ' '))
                ws.cell(row=row, column=2, value=md[1])
                ws.cell(row=row, column=3, value=md[2])
                ws.cell(row=row, column=4, value="")
                ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
                apply_border_to_row(ws, row, 1, 4); row += 1
        row += 1

        if report.inspection_type != "なし" and report.inspector:
            ws[f'A{row}'] = "【社内検査】"
            ws[f'A{row}'].font = Font(bold=True, size=11, color="E91E63")
            row += 1
            ws.cell(row=row, column=1, value="検査種類").fill = fill_gray; ws.cell(row=row, column=2, value="検査者").fill = fill_gray; ws.cell(row=row, column=3, value="判定").fill = fill_gray
            apply_border_to_row(ws, row, 1, 4); row += 1
            ws.cell(row=row, column=1, value=report.inspection_type); ws.cell(row=row, column=2, value=report.inspector); ws.cell(row=row, column=3, value=report.judge)
            apply_border_to_row(ws, row, 1, 4); row += 2

        text_sections = [
            ("明日の作業内容", report.tomorrow_work),
            ("顧客からの指示事項", report.customer),
            ("社内連絡", report.internal),
            ("安全指示・確認事項", report.safety),
            ("作業所長巡視所見処置・確認事項", f"AM ({report.am_time}): 異常{report.am_abnormal}\n{report.am_text}\n\nPM ({report.pm_time}): 異常{report.pm_abnormal}\n{report.pm_text}")
        ]

        for title, content in text_sections:
            ws.merge_cells(f'A{row}:D{row}')
            ws[f'A{row}'] = title
            ws[f'A{row}'].font = font_bold
            ws[f'A{row}'].fill = fill_gray
            ws[f'A{row}'].alignment = align_left
            apply_border_to_row(ws, row, 1, 4)
            row += 1

            ws.merge_cells(f'A{row}:D{row}')
            ws[f'A{row}'] = content
            ws[f'A{row}'].alignment = align_top_left

            line_breaks = str(content).count('\n') + 1
            ws.row_dimensions[row].height = max(45, line_breaks * 18)
            apply_border_to_row(ws, row, 1, 4)
            row += 2

    if not wb.sheetnames:
        wb.create_sheet(title="データなし")

    stream = io.BytesIO()
    wb.save(stream); stream.seek(0)

    filename = f"Report_{req.project_no}_{req.period}.xlsx"
    encoded_filename = urllib.parse.quote(filename)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 10000))
    uvicorn.run(app, host="0.0.0.0", port=port)
